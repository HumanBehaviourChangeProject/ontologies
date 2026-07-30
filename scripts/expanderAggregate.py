from enum import Enum
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Font
import sys
import argparse

# --------------------------------------------------------------------------- #
# ID handling
# --------------------------------------------------------------------------- #
# Every generated entity gets a stable BCIO id drawn from this reserved space.
# The invariants we must preserve across re-runs (see IdAllocator):
#   * an id that has ever been handed out is NEVER handed out again, and
#   * an entity that is no longer generated is NOT deleted but marked Obsolete.
ID_SPACE = (15000, 16000)
PREFIX = 'BCIO'
ID_WIDTH = 6

OBSOLETE_STATUS = "Obsolete"

# Column that is only ever populated on rows this script generates. We use its
# presence to tell "generated" rows apart from hand-curated input rows when we
# read a previously generated output file back in.
GENERATED_MARKER_COL = "REL 'aggregate of'"


class Kind(Enum):
    NUMBER = "number"
    VALUE = "value"
    PEOPLE = "people"
    ATTRIBUTES = "attributes"
    ROLES = "roles"
    PAST_BEHAVIOUR = "past behaviour"


class Aggregate(Enum):
    MEAN = "mean"
    MINIMUM = "minimum"
    MAXIMUM = "maximum"
    MEDIAN = "median"
    PERCENTAGE = "percentage"
    PROPORTION = "proportion"


def get_aggregate_definition(statistic: str, aggregate: Aggregate, kind: Kind) -> str:
    definition = {
        (Aggregate.MEAN, Kind.NUMBER): f"A(n) {statistic} population statistic that is the mean number of {statistic} in the population.",
        (Aggregate.MINIMUM, Kind.NUMBER): f"A(n) {statistic} population statistic that is the minimum number of {statistic} in the population.",
        (Aggregate.MAXIMUM, Kind.NUMBER): f"A(n) {statistic} population statistic that is the maximum number of {statistic} in the population.",
        (Aggregate.MEDIAN, Kind.NUMBER): f"A(n) {statistic} population statistic that is the median number of {statistic} in the population.",

        (Aggregate.MEAN, Kind.VALUE): f"A(n) {statistic} population statistic that is the mean value of {statistic} in the population.",
        (Aggregate.MINIMUM, Kind.VALUE): f"A(n) {statistic} population statistic that is the minimum value of {statistic} in the population.",
        (Aggregate.MAXIMUM, Kind.VALUE): f"A(n) {statistic} population statistic that is the maximum value of {statistic} in the population.",
        (Aggregate.MEDIAN, Kind.VALUE): f"A(n) {statistic} population statistic that is the median value of {statistic} in the population.",
        (Aggregate.PERCENTAGE, Kind.VALUE): f"A(n) {statistic} population statistic that is the percentage value of {statistic} in the population.",
        (Aggregate.PROPORTION, Kind.VALUE): f"A(n) {statistic} population statistic that is the proportion of individuals having a {statistic} in the population.",

        (Aggregate.PERCENTAGE, Kind.PEOPLE): f"A(n) {statistic} population statistic that is the percentage of people that are a {statistic} in the population.",
        (Aggregate.PROPORTION, Kind.PEOPLE): f"A(n) {statistic} population statistic that is the proportion of people that are a {statistic} in the population.",

        (Aggregate.PERCENTAGE, Kind.ATTRIBUTES): f"A(n) {statistic} population statistic that is the percentage of people that are {statistic} in the population.",
        (Aggregate.PROPORTION, Kind.ATTRIBUTES): f"A(n) {statistic} population statistic that is the proportion of people that are {statistic} in the population.",

        (Aggregate.PERCENTAGE, Kind.ROLES): f"A(n) {statistic} population statistic that is the percentage of people that have a {statistic} in the population.",
        (Aggregate.PROPORTION, Kind.ROLES): f"A(n) {statistic} population statistic that is the proportion of people that have a {statistic} in the population.",

        (Aggregate.PERCENTAGE, Kind.PAST_BEHAVIOUR): f"A(n) {statistic} population statistic that is the percentage of people that have {statistic} in the population.",
        (Aggregate.PROPORTION, Kind.PAST_BEHAVIOUR): f"A(n) {statistic} population statistic that is the proportion of people that have {statistic} in the population.",

    }.get((aggregate, kind))

    if definition is None:
        raise ValueError(
            f"Unknown aggregate {aggregate} for kind {kind} in statistic '{statistic}'"
        )

    return definition


class IdAllocator:
    """Hands out BCIO ids while guaranteeing they are never reused.

    An id is "reserved" as soon as it is seen anywhere (the input file or a
    previously generated output file). Reserved ids are removed from the pool of
    free ids, so :meth:`new` can never return an id that has already been used
    for something else -- even for entities that have since been deleted.
    """

    def __init__(self, id_space: Tuple[int, int], prefix: str, width: int):
        self.prefix = prefix
        self.width = width
        # candidate ids, kept ordered so allocation is deterministic
        self._free: List[int] = list(range(id_space[0] + 1, id_space[1]))
        self._free_set = set(self._free)
        self._reserved: set = set()

    @staticmethod
    def to_int(id_val) -> Optional[int]:
        """Parse the numeric part of an id such as ``BCIO:015123`` or ``15123``."""
        if id_val is None:
            return None
        if isinstance(id_val, int):
            return id_val
        s = str(id_val).strip()
        if s == "":
            return None
        if ":" in s:
            s = s.split(":", 1)[1].strip()
        try:
            return int(s)
        except ValueError:
            return None

    def reserve(self, id_val) -> None:
        """Mark an id as used so it can never be allocated again."""
        n = self.to_int(id_val)
        if n is None:
            return
        self._reserved.add(n)
        if n in self._free_set:
            self._free_set.discard(n)
            # keep list in sync (linear removal is fine for a ~1000 id space)
            try:
                self._free.remove(n)
            except ValueError:
                pass

    def new(self) -> str:
        if not self._free:
            raise RuntimeError(
                f"Ran out of free ids in space {ID_SPACE}. Widen ID_SPACE."
            )
        n = self._free.pop(0)
        self._free_set.discard(n)
        self._reserved.add(n)
        return f"{self.prefix}:{str(n).zfill(self.width)}"


def is_obsolete(row: Dict[str, object]) -> bool:
    status = row.get("Curation status")
    return status is not None and str(status).strip() == OBSOLETE_STATUS


def is_generated(row: Dict[str, object]) -> bool:
    """True for rows this script produced (they always set the marker column)."""
    marker = row.get(GENERATED_MARKER_COL)
    return marker is not None and str(marker).strip() != ""


def load_prior_output(path: Path):
    """Read a previously generated ``*_Expanded.xlsx`` file, if it exists.

    Returns ``(header, prior_rows, generated_active)`` where ``prior_rows`` is
    every non-empty row (as a dict) from the previous output -- the complete
    record of ids that have ever been handed out -- and ``generated_active``
    maps a currently active (non-obsolete) *generated* row's Label -> its row
    dict, used to keep statistic ids stable across re-runs.
    """
    if not path.exists():
        return None, [], {}

    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    header = [c.value for c in next(sheet.rows)]

    prior_rows: List[Dict[str, object]] = []
    generated_active: Dict[str, Dict[str, object]] = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = dict(zip(header, row))
        if not any(v is not None and str(v).strip() != "" for v in values.values()):
            continue
        prior_rows.append(values)
        if is_generated(values) and not is_obsolete(values):
            label = values.get("Label")
            if label is not None and str(label).strip() != "":
                generated_active[str(label)] = values

    return header, prior_rows, generated_active


def build_extra_rows(
    header,
    row,
    aggregate: str,
    kind: Kind,
    parents: Dict[str, str],
    label_to_id: Dict[str, str],
    allocator: IdAllocator,
) -> List[Dict[str, object]]:
    """Generate the population-statistic rows for a single source row.

    Ids are assigned by Label through ``label_to_id`` so that re-running the
    script yields the *same* id for the same statistic; only genuinely new
    statistics draw a fresh id from ``allocator``.
    """
    aggregate = aggregate.lower()
    aggregate_list = aggregate.split(";")
    # the base "aggregate" (parent) statistic always comes first
    aggregate_list.insert(0, "aggregate")

    extra_rows = []
    for agg in aggregate_list:
        agg = agg.strip()
        extra_values: Dict[str, object] = {}
        name = ""
        for key, cell in zip(header, row):
            if key == "Label":
                if agg == "aggregate":
                    extra_values[key] = f"{cell.value} population statistic"
                else:
                    extra_values[key] = f"{agg} {cell.value} population statistic"
                name = str(cell.value)
            elif key == "Parent":
                if agg == "aggregate":
                    parent = parents.get(name, None)
                    if parent is not None and parent != "":
                        extra_values[key] = parent + " population statistic"
                    else:
                        extra_values[key] = "population statistic"
                else:
                    extra_values[key] = f"{name} population statistic"
            elif key == "ID":
                # assigned after the Label is known (see below)
                extra_values[key] = None
            elif key == "Definition":
                if agg == "aggregate":
                    extra_values[key] = f"A population statistic about {name}."
                else:
                    extra_values[key] = get_aggregate_definition(
                        name, Aggregate(agg.lower().strip()), kind
                    )
            elif key in ["Curation status"]:
                extra_values[key] = str(cell.value) if cell.value != "External" else "Published"
            elif key in ["Sub-ontology"]:
                extra_values[key] = str(cell.value)
            elif key == GENERATED_MARKER_COL:  # REL 'aggregate of'
                extra_values[key] = name
            else:
                extra_values[key] = ""

        if not any(v not in (None, "") for v in extra_values.values()):
            continue

        # Stable id assignment keyed on the (unique) generated Label.
        if "ID" in extra_values:
            label = str(extra_values.get("Label", "")).strip()
            existing = label_to_id.get(label)
            if existing is not None:
                extra_values["ID"] = existing
            else:
                new_id = allocator.new()
                extra_values["ID"] = new_id
                label_to_id[label] = new_id

        extra_rows.append(extra_values)

    return extra_rows


## PROGRAM EXECUTION --- required argument: input file name
if __name__ == '__main__':

    parser = argparse.ArgumentParser()
    parser.add_argument('--inputExcel', '-i', help='Name of the input Excel spreadsheet file')

    args = parser.parse_args()

    inputFileName = args.inputExcel

    if inputFileName is None:
        parser.print_help()
        sys.exit('Not enough arguments. Expected at least -i "Excel file name" ')

    pathpath = str(Path(inputFileName).parents[0])
    basename = str(Path(inputFileName).stem)
    suffix = str(Path(inputFileName).suffix)

    output_path = Path(pathpath) / f"{basename}_Expanded.xlsx"

    wb: openpyxl.Workbook = openpyxl.load_workbook(inputFileName)
    sheet = wb.active
    assert sheet is not None
    header = [i.value for i in next(sheet.rows)]

    kind_index = header.index("Kind") if "Kind" in header else None
    assert kind_index is not None, "Header must contain 'Kind' column"

    allocator = IdAllocator(ID_SPACE, PREFIX, ID_WIDTH)

    # ------------------------------------------------------------------ #
    # 1. Reserve every id that already exists in the *input* file.
    # ------------------------------------------------------------------ #
    for row in sheet[2:sheet.max_row]:
        for key, cell in zip(header, row):
            if key == "ID" and cell.value is not None:
                allocator.reserve(cell.value)

    # ------------------------------------------------------------------ #
    # 2. Load the previously generated output. It is the persistent record
    #    of which ids have been handed out and which statistics existed, so
    #    we can keep ids stable and never delete anything.
    # ------------------------------------------------------------------ #
    prior_header, prior_rows, prior_active = load_prior_output(output_path)

    # Reserve EVERY id ever written to the output (generated statistics *and*
    # copied original entities) so a deleted entity's id can never be reused.
    for prow in prior_rows:
        allocator.reserve(prow.get("ID"))

    # Stable Label -> id map for the currently active generated statistics.
    label_to_id: Dict[str, str] = {}
    for label, prow in prior_active.items():
        pid = prow.get("ID")
        if pid is not None and str(pid).strip() != "":
            label_to_id[label] = str(pid)

    # Column order: honour the existing output file so re-runs produce minimal
    # diffs; fall back to the input order on first generation.
    output_header = prior_header if prior_header is not None else header

    # ------------------------------------------------------------------ #
    # 3. Build the parents dict (label -> parent) used to nest the base
    #    "population statistic" entities.
    # ------------------------------------------------------------------ #
    entries: Dict[str, Tuple[str, bool]] = {}
    for row in sheet[2:sheet.max_row]:
        values = dict(zip(header, [v.value for v in row]))
        label = values["Label"].strip() if values["Label"] is not None else ""
        if label == "":
            continue
        entries[label] = (
            values["Parent"].strip() if values["Parent"] is not None else "",
            values["Aggregate"] is not None and str(values["Aggregate"]).strip() != ""
        )

    parents = {k: p for k, (p, _) in entries.items() if entries.get(p, ("", False))[1] is True}

    # ------------------------------------------------------------------ #
    # 4. Copy the original input rows verbatim.
    # ------------------------------------------------------------------ #
    rows: List[Dict[str, object]] = []
    for row in sheet[2:sheet.max_row]:
        values = {key: cell.value for key, cell in zip(header, row)}
        if any(values.values()):
            rows.append(values)

    # ------------------------------------------------------------------ #
    # 5. Generate the population-statistic rows. Obsolete source entities do
    #    not spawn (new) statistics -- their previously generated statistics
    #    are handled by the carry-forward step below.
    # ------------------------------------------------------------------ #
    current_generated_labels: set = set()
    for i, row in enumerate(sheet[2:sheet.max_row]):
        try:
            values = {key: cell.value for key, cell in zip(header, row)}
            if is_obsolete(values):
                continue

            aggregate_cell = values.get("Aggregate")
            if aggregate_cell is None or str(aggregate_cell).strip() == "":
                continue

            kind_val = row[kind_index].value if kind_index is not None else None
            if kind_val is None or str(kind_val).strip() == "":
                print("No kind found for row:", values.get("ID"), values.get("Label"))
                continue
            kind = Kind(str(kind_val).strip())

            extra_rows = build_extra_rows(
                header, row, str(aggregate_cell), kind, parents, label_to_id, allocator
            )
            for extra_row in extra_rows:
                rows.append(extra_row)
                lbl = str(extra_row.get("Label", "")).strip()
                if lbl:
                    current_generated_labels.add(lbl)
        except Exception as e:
            print(f"Error processing row {i}")
            raise e

    # ------------------------------------------------------------------ #
    # 6. Carry forward everything from the previous output whose id is no
    #    longer present: a statistic no longer generated, or an original
    #    entity removed from the input. Such rows are never deleted -- they are
    #    retained with Curation status "Obsolete" (already-obsolete rows keep
    #    that status), so no id ever disappears from the file.
    # ------------------------------------------------------------------ #
    current_ids = {
        IdAllocator.to_int(r.get("ID"))
        for r in rows
        if IdAllocator.to_int(r.get("ID")) is not None
    }

    newly_obsoleted = 0
    still_obsolete = 0
    for prow in prior_rows:
        pid = IdAllocator.to_int(prow.get("ID"))
        if pid is None or pid in current_ids:
            continue  # id still exists in the freshly built output
        carried = dict(prow)
        if is_obsolete(carried):
            still_obsolete += 1
        else:
            carried["Curation status"] = OBSOLETE_STATUS
            print(f"Carrying forward {carried.get('ID')} as Obsolete")
            newly_obsoleted += 1
        rows.append(carried)

    print(
        f"Generated {len(current_generated_labels)} active statistics; "
        f"{newly_obsoleted} entities newly obsoleted; "
        f"{still_obsolete} already-obsolete entities carried forward."
    )

    # ------------------------------------------------------------------ #
    # 7. Write the output using the stable column order.
    # ------------------------------------------------------------------ #
    save_wb = openpyxl.Workbook()
    save_sheet = save_wb.active

    for c, col_name in enumerate(output_header):
        save_sheet.cell(row=1, column=c + 1).value = col_name
        save_sheet.cell(row=1, column=c + 1).font = Font(size=12, bold=True)

    for r, values in enumerate(rows):
        for c, col_name in enumerate(output_header):
            val = values.get(col_name, "")
            save_sheet.cell(row=r + 2, column=c + 1).value = val

    save_wb.save(str(output_path))
    print(f"success -> {output_path}")
