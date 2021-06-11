import os
import csv
import re
from openpyxl import Workbook, load_workbook

prjdir = '/Users/hastingj/Work/Onto/HBCP/ontologies/MoA/'

os.chdir(prjdir)

sub_dir = "inputs"
file_name = "MoA.xlsx"

id_root = 6001 # BCI MoA is BCIO:006000

wb = load_workbook(sub_dir+'/'+file_name)
sheet = wb.active

sheet_data = sheet.rows
header = [i.value for i in next(sheet_data)]

#header = [i.value for i in next(sheet_data)]
#print (header)

data_out = {}
data_out["ID"] = []
data_out["Label"] = []
data_out["Parent class"] = []
data_out["Definition"] = []
data_out["Cross-reference"] = []
data_out["Synonyms"] = []
data_out["Elaborations"] = []
data_out["Examples"] = []

data = { 'last_ul' : '', 'last_sl1' : '', 'last_sl2' : '', 'last_sl3' : '', 'last_sl4' : '' , 'last_sl5' : '', 'last_sl6':'', 'id' : id_root }

for row in sheet_data:
    row_data = [i.value for i in row]
    if any(row_data):
        ul = row_data[header.index('Level 1')]
        sl1 = row_data[header.index('Level 2')]
        sl2 = row_data[header.index('Level 3')]
        sl3 = row_data[header.index('Level 4')]
        sl4 = row_data[header.index('Level 5')]
        sl5 = row_data[header.index('Level 6')]
        sl6 = row_data[header.index('Level 7')]
        sl7 = row_data[header.index('Level 8')]
        dfn = row_data[header.index('Definitions')]
        if dfn:
            dfn = dfn.replace('\n', ' ').replace('\r', '').replace('\"','\'')
        #uri = row_data[header.index('URI')]
        syn = row_data[header.index('Synonym')]
        elab = row_data[header.index('Elaborations')]
        exm = row_data[header.index('Example')]

        #if uri and "http://purl.obolibrary.org/" in uri:
        #    classId = uri.replace("http://purl.obolibrary.org/obo/","").replace("_",":")
        #else:
        classId = "BCIO:00"+str(data['id'])
        data['id'] = data['id']+1
        data_out["ID"].append(classId)
        data_out["Definition"].append(dfn)
        #data_out["Cross-reference"].append(uri)
        data_out["Synonyms"].append(syn)
        data_out["Elaborations"].append(elab)
        data_out["Examples"].append(exm)

        if (ul and len(ul)>0):
            # New upper level term
            label = re.sub(r"^\d*\. ", "", ul)
            data_out["Label"].append(label)
            data_out["Parent class"].append("Thing")
            data['last_ul'] = label
        if (sl1 and len(sl1)>0):
            label = re.sub(r"^\d*\. ", "", sl1)
            data_out["Label"].append(label)
            data_out["Parent class"].append(data['last_ul'])
            data['last_sl1'] = label
        if (sl2 and len(sl2)>0):
            label = re.sub(r"^\d*\. ", "", sl2)
            data_out["Label"].append(label)
            data_out["Parent class"].append(data['last_sl1'])
            data['last_sl2'] = label
        if (sl3 and len(sl3)>0):
            label = re.sub(r"^\d*\. ", "", sl3)
            data_out["Label"].append(label)
            data_out["Parent class"].append(data['last_sl2'])
            data['last_sl3'] = label
        if (sl4 and len(sl4)>0):
            label = re.sub(r"^\d*\. ", "", sl4)
            data_out["Label"].append(label)
            data_out["Parent class"].append(data['last_sl3'])
            data['last_sl4'] = label
        if (sl5 and len(sl5)>0):
            label = re.sub(r"^\d*\. ", "", sl5)
            data_out["Label"].append(label)
            data_out["Parent class"].append(data['last_sl4'])
            data['last_sl5'] = label
        if (sl6 and len(sl6)>0):
            label = re.sub(r"^\d*\. ", "", sl6)
            data_out["Label"].append(label)
            data_out["Parent class"].append(data['last_sl6'])
            data['last_sl6'] = label
        if (sl7 and len(sl7)>0):
            label = re.sub(r"^\d*\. ", "", sl7)
            data_out["Label"].append(label)
            data_out["Parent class"].append(data['last_sl6'])

# Save output to file
wb_out = Workbook()
sheet_out = wb_out.active

# Write header:
for k,i in zip( data_out.keys(), range(len(data_out.keys())) ):
    sheet_out.cell(row=1,column=i+1).value =  k

# Write row values:
for i in range(len(data_out['ID'])):
    for k,j in zip( data_out.keys(), range(len(data_out.keys())) ):
        if len(data_out[k])>i:
            sheet_out.cell(row=(i+2),column=j+1).value = data_out[k][i]

wb_out.save("MOA-converted.xlsx")








