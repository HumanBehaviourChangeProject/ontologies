
import os
import openpyxl
from collections import Counter
import re
import argparse
import sys

dirpath = 'Work/Onto/HBCP/ontologies/'
inputfile = 'MechanismOfAction/bcio_moa.xlsx'
externalfile = 'Upper Level BCIO/inputs/BCIO_External_Imports.xlsx'

external_tmp = []
external_list = []
id_list = []
label_list = []

externalFileName = dirpath + externalfile
wb = openpyxl.load_workbook(externalFileName) 
sheet = wb.active
data = sheet.rows
rows = []
header = [i.value for i in next(data)]

for row in sheet[2:sheet.max_row]:
    for key, cell in zip(header, row):
        if key == "IDs" and cell.value != None:
            external_tmp.append(re.findall(r'\[.*?\]', cell.value))
            external_list = [item for sublist in external_tmp for item in sublist] #flatten list
            external_list = [item.strip('[]') for item in external_list]
            external_list = [item.replace('\ufeff', '') for item in external_list] #remove special character #todo: is there a better way to sanitize?
# print("external_list is: ", external_list)

files = [dirpath+inputfile]

allInfo = []
for f in files:
    if f.endswith("xlsx"):
        inputFileName = f
        wb = openpyxl.load_workbook(inputFileName)
        sheet = wb.active
        data = sheet.rows
        rows = []
        header = [i.value for i in next(data)]

        for row in sheet[2:sheet.max_row]:
            values = {}
            values["Sheet"] = f
            for key, cell in zip(header, row):
                if key == "ID" and cell.value != None:
                    id_list.append(cell.value)
                    values[key] = cell.value
                if key == "Label" and cell.value != None:
                    values[key] = cell.value
            allInfo.append(values)

# compare external_list to id_list and return duplicates:
returnList = []
present = False
for d in allInfo:
    instance = {}
    for k, v in d.items():
        if k == "ID": 
            if "BCIO" in v:
                pass
            else: 
                for i in external_list:
                    if v == i and v != None and v.strip() != '':
                        present = True                    
                
                if not present:                    
                    if v.strip() != '': # check blank ID
                        # instance.update(d) # all at once  
                        instance[k] = v
                        instance["Label"] = d["Label"]
                        instance["Sheet"] = d["Sheet"] 
                        returnList.append(instance)

listsByOnto = {}

print("ID's not present in External_Imports: ")  
print("")
count = 0;
for r in returnList:
    count = count + 1
    print(count, ": ", r)  
    print("")   
    ontoname = r['ID'].split(":")[0]
    if ontoname not in listsByOnto:
        listsByOnto[ontoname] = set()
    listsByOnto[ontoname].add(f"{r['Label']} [{r['ID']}]")
    
            
for k,v in listsByOnto.items():
    print(k)
    print(*v,sep=';')      
    print('\n') 
            
