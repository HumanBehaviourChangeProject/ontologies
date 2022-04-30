import os
from pathlib import Path
import re
import openpyxl
from openpyxl.styles import Font
import csv
import codecs
import sys
import argparse
import subprocess
import io

def generate_unique_id(base_id, id_list, num):
    num_list = []
    split_base = base_id.split(":")
    split_base_num = int(split_base[1]) + num

    for id in id_list:
        split_id = id.split(":")
        split_id_num = int(split_id[1])
        num_list.append(split_id_num)
    num_list.sort()
    if split_base_num in num_list:
        return split_id[0] + ":" + str(num_list[-1]).zfill(6)
    else: 
        return split_id[0] + ":" + str(split_base_num).zfill(6)

def add_extra_values(header, row, aggregate, id_list):
    aggregate_list = aggregate.split(";")
    extra_rows = []
    i = 0
    for agg in aggregate_list:
        i = i+1
        extra_values = {}
        name = ""
        for key, cell in zip(header, row):
            if key == "Label":
                extra_values[key] = agg + " " + str(cell.value)
                name = str(cell.value)
            elif key == "Parent":
                extra_values[key] = name 
            elif key == "ID": 
                new_id = generate_unique_id(cell.value, id_list, i)
                extra_values[key] = new_id 
            elif key == "Definition":
                extra_values[key] = "The " + agg + " of " + name
            else:
                extra_values[key] = "" 
        if any(extra_values.values()):
            extra_rows.append(extra_values)
    return extra_rows


## PROGRAM EXECUTION --- required argument: input file name
if __name__ == '__main__':

    parser=argparse.ArgumentParser()
    parser.add_argument('--inputExcel', '-i',help='Name of the input Excel spreadsheet file')
    
    args=parser.parse_args()

    inputFileName = args.inputExcel
    
    if inputFileName is None :
        parser.print_help()
        sys.exit('Not enough arguments. Expected at least -i "Excel file name" ')

    pathpath = str(Path(inputFileName).parents[0])
    basename = str(Path(inputFileName).stem)
    suffix = str(Path(inputFileName).suffix)
    
    wb = openpyxl.load_workbook(inputFileName) 
    sheet = wb.active
    data = sheet.rows
    rows = []
    aggregate_list = ["Mean", "Minimum", "Maximum", "Median"]
    header = [i.value for i in next(data)]
    
    #build ID list:
    id_list = []
    for row in sheet[2:sheet.max_row]:
        values = {}
        extra_rows = []
        for key, cell in zip(header, row):
            values[key] = cell.value
            if key == "ID" and cell.value != None:
                id_list.append(cell.value)

    for row in sheet[2:sheet.max_row]:
        values = {}
        extra_rows = []
        for key, cell in zip(header, row):
            values[key] = cell.value
            if key == "Aggregate" and cell.value != None:
                extra_rows = add_extra_values(header, row, cell.value, id_list)
        if any(values.values()):
            rows.append(values)
        for extra_row in extra_rows:
            rows.append(extra_row)
        
    for r in range(len(rows)):
        row = [v for v in rows[r].values()]
        if "Aggregate" in header:
            cell = row[header.index("Aggregate")]    

    #new sheet to save:

    save_wb = openpyxl.Workbook()
    save_sheet = save_wb.active

    for c in range(len(header)):
        save_sheet.cell(row=1, column=c+1).value=header[c]
        save_sheet.cell(row=1, column=c+1).font = Font(size=12,bold=True)
    for r in range(len(rows)):
        row = [v for v in rows[r].values()]
        for c in range(len(header)):
            save_sheet.cell(row=r+2, column=c+1).value=row[c]
            
    #save:   
    save_wb.save(pathpath + "/" + basename + "_Expanded.xlsx")
    print("success")
    

